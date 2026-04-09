import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 新数据
data = [
    ["序号", "姓名", "学号", "号码"],
    [1, "郑力凡", "S325517174", 77],
    [2, "卢宏伟", "S325517083", 4],
    [3, "张功琦", "S325517153", 0],
    [4, "孙雪恒", "S325517103", 11],
    [5, "严佳俊", "S325517142", 16],
    [6, "莫希华", "S325517088", 20],
    [7, "姜峰", "S324517047", 45],
    [8, "万睿", "S325517108", 8],
    [9, "武烨朝", "S325517128", 29],
    [10, "庞云鹤", "B224510008", 17],
    [11, "刘鹏财", "S325517076", 10],
    [12, "朱亦琪", "S325517181", 28],
    [13, "李文", "B220080004", 7],
]

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "电子队报名"

# 写入数据
for row_idx, row_data in enumerate(data, 1):
    for col_idx, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # 设置表头样式
        if row_idx == 1:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

# 调整列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 10

# 保存到桌面
wb.save(r"C:\Users\hp\Desktop\电子队报名.xlsx")
print("文件已保存到桌面：电子队报名.xlsx")
